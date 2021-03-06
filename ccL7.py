import openpyxl
import random


def gen_l7(configList,path):
    global serviceDict
    global max_entry_id
    global serviceEntryIdDict
    serviceDict = {}
    serviceEntryIdDict = {}
    max_entry_id = -1
    commandList = []
    # print("把excel表拖入cmd窗口\n")
    # excel_path = input()
    # excel_path = "C:\L7chargingcontext.xlsx"
    excel_path = path+"\\内容计费整理L7.xlsx"
    # print("把内容计费的配置log表拖入cmd窗口\n")
    # chargingContextLog_path = input()
    # chargingContextLog_path = "D:\chargingContext20180403\sae133-config-20180330.txt"
    # configFile = open(chargingContextLog_path, 'r')
    # configList = configFile.readlines()
    # configFile.close()

    excel = openpyxl.load_workbook(excel_path)
    sheet = excel["L7"]
    serviceList = []
    # 该函数会根据分割符来把一条条目中包含port range的分成若干条
    serviceList = getServiceListByList(sheet, 3)
    # print(serviceList)
    resultList = []
    resultList = arrangeTheList(serviceList)
    # print(resultList)

    resultList = arrangeTheList_2(resultList)
    # print(resultList)

    allEntryIdList = []
    getAllEntryIdList(allEntryIdList, configList)
    # print("所有entry id:"+str(allEntryIdList))

    for resultlst in resultList:

        commandList.append(resultlst[0][3] + "业务进行增删操作\n")
        # commandList.append("exit all\n")
        # commandList.append("configure application-assurance group 1:1 policy\n")
        # commandList.append("begin\n")
        # commandList.append("app-filter\n")
        setPRUCRUtoServiceDict(resultlst[0], configList)

        dnsMatchList = []
        # dnsMatchList = getNDSMatchListByConfigureLog(resultlst,configList)
        addNDSMatch2CommandList(commandList, dnsMatchList)
        addEntryIdtoserviceEntryIdDict(resultlst[0][3], configList)
        for tupline in resultlst:
            if tupline[1] == "新增":
                entryId = getTheCompatibleEntryIdByDict(tupline)
                addTheCommandtoList_Entry(commandList, tupline, entryId)
            else:
                # 删除entry

                # 获取该业务的所有entry存于字典中,用来判断该业务是否删干净

                deleteEntryId = getTheDeleteEntryId(tupline, configList)
                DeleteTheEntry(commandList, tupline, deleteEntryId)
                LimitTheServiceSpeed(commandList, tupline)

                # break

            PR_PRU_CRU_Process(commandList, resultlst[0], configList)
            PR_PRU_CRU_Delete(commandList, resultlst[0], configList)

        commandList.append("\n\n")

    #print("该业务所有PRU,CRU,PR的配置情况")
    config_stats=[]
    config_name=[]
    # for key in serviceDict:
    #     config_name.append(key)
    #     config_stats.append(serviceDict[key])

    #print("该业务的所有ID")
    #for key in serviceEntryIdDict:
        #print(key, serviceEntryIdDict[key])

    fo = open(path+"\\testL7.txt", "w")
    fo.writelines(commandList)
    fo.close()

    return serviceDict
    # exit(7)


def getServiceListByList(sheet, startRow):
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    # firstLineServiceId = sheet.cell(row=startRow, column=serviceId_col).value
    retList = []

    for rowNumber in range(startRow, sheet.max_row + 1):
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value
        layerLag = "L7"
        portsplit_1 = "|"
        portsplit_2 = "=="
        portL4List = []
        portstr = portNumberL4
        if portNumberL4 != None:
            if portsplit_2 in str(portNumberL4):
                portstr = portNumberL4.split(portsplit_2)[0]
                portL4List.append(portNumberL4.split(portsplit_2)[1])
            else:
                pass
            if portsplit_1 in str(portstr):
                portstr = portstr.split(portsplit_1)
                for p in portstr:
                    portL4List.append(p)
            else:
                portL4List.append(portstr)
            # print(portL4List)
            for p in portL4List:
                retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, p, urlL7))
        else:
            retList.append(
                (layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
    retList = list(set(retList))

    return retList


def arrangeTheList(lst):
    sList = []
    retList = []
    tempList = []
    for tup in lst:
        layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
        sList.append(serviceId)
    sList = list(set(sList))
    for sValue in sList:
        for tup in lst:
            layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []
    return retList


def arrangeTheList_2(rtList):
    new_ret_list = []
    for lst in rtList:
        tmplist = []
        for tup in lst:
            if tup[1] == "新增":
                tmplist.append(tup)
        for tup in lst:
            if tup[1] == "删除":
                tmplist.append(tup)
        new_ret_list.append(tmplist)
    return new_ret_list


def getNDSMatchListByConfigureLog():
    pass


def addNDSMatch2CommandList(comdLst, dnsMaLst):
    pass


def getTheCompatibleEntryIdByDict(tup):
    global serviceEntryIdDict
    global allEntryIdList
    retId = -1
    retId = 30000

    return retId


def addTheCommandtoList_Entry(comLst, tup, enId):
    # print(tup)
    layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
    comLst.append("exit all\n")
    comLst.append("configure application-assurance group 1:1 policy\n")
    comLst.append("begin\n")
    comLst.append("app-filter\n")
    comLst.append("entry " + str(enId) + " create\n")
    expression_1 = ""
    expression_2 = ""
    http_port = ""
    if url != None:
        url = url.replace("http://", "").replace("https://", "")
        if url[0] != "*":
            url = "^" + url
        if url[-1] != "*":
            url = url + "$"

        prefix = ""
        suffix = ""
        if "/*" in url:
            suffix = "/*"
            url = url.replace("/*", "")
            url = url + "$"
            expression_2 = "^/*"

        expression_1 = url
        if ":*" not in url and ":" in url:
            expression_1 = url.split(":")[0] + "$"
            expression_2 = "^" + url.split(":")[1][url.split(":")[1].index("/"):len(url.split(":")[1])].replace("$",
                                                                                                                "") + "/*"
            http_port = url.split(":")[1][0:url.split(":")[1].index("/")]

        comLst.append('expression 1 http-host eq "' + expression_1 + '"\n')
    if expression_2 != "":
        comLst.append('expression 2 http-uri eq "' + expression_2 + '"\n')
    if ipAddress == None:
        comLst.append("server-address eq 10.0.0.172/32\n")
    else:
        if "/" not in ipAddress:
            ipAddress = ipAddress + "/32"
        comLst.append('server-address eq ' + ipAddress + '\n')
    if portNumber != None:
        if " " in str(portNumber):
            comLst.append('server-port range ' + str(portNumber) + '\n')
        else:
            comLst.append('server-port eq ' + str(portNumber) + '\n')
    if http_port != "":
        comLst.append('http-port eq ' + str(http_port) + '\n')
    comLst.append('application "APP_' + serviceName + '"\n')
    comLst.append("no shutdown\n")
    comLst.append("exit\n")
    comLst.append("\n")
    # 纯7L的地址（网址应该创建dns-catch）
    global max_entry_id
    # if ipAddress == None and portNumber == None and tup[7].upper() != tup[7].lower():
    if ipAddress == None and portNumber == None and tup[7] != None:
        max_entry_id += 1
        comLst.append("exit all\n")
        comLst.append("configure application-assurance group 1:1 policy\n")
        comLst.append("begin\n")
        comLst.append("app-filter\n")
        comLst.append("entry " + str(enId + 1) + " create\n")
        if url != None:
            url = url.replace("http://", "").replace("https://", "").replace("^", "")
            if url[0] != "*":
                url = "^" + url
            if url[-1] != "*":
                # print(url)
                url = url.replace("$", "") + "$"

            prefix = ""
            suffix = ""
            if "/*" in url:
                suffix = "/*"
                url = url.replace("/*", "")
                # url = url + "$"
                expression_2 = "^/*"

            expression_1 = url
            # print(expression_1)
            if ":*" not in url and ":" in url:
                expression_1 = url.split(":")[0] + "$"
                expression_2 = "^" + url.split(":")[1][url.split(":")[1].index("/"):len(url.split(":")[1])].replace("$",
                                                                                                                    "") + "/*"
                http_port = url.split(":")[1][0:url.split(":")[1].index("/")]

            comLst.append('expression 1 http-host eq"' + expression_1 + '"\n')
        if expression_2 != "":
            comLst.append('expression 2 http-uri eq "' + expression_2 + '"\n')

        comLst.append('server-address eq dns-ip-cache "TrustedCache"\n')
        if portNumber != None:
            if " " in str(portNumber):
                comLst.append('server-port range ' + str(portNumber) + '\n')
            else:
                comLst.append('server-port eq ' + str(portNumber) + '\n')
        if http_port != "":
            comLst.append('http-port eq ' + str(http_port) + '\n')
        comLst.append('application "APP_' + serviceName + '"\n')
        comLst.append("no shutdown\n")
        comLst.append("exit\n")
        comLst.append("\n\n")
        # 因为要添加dns-catch得有两entry所以得添加2次,这里先添加一次
    serviceEntryIdDict[tup[3]].append(enId)


def getServiceEntryList(serviceName, cLst):
    retLst = []
    tmpLst = []
    # print("servicename is "+ serviceName)
    for i in range(0, len(cLst)):
        if 'application "APP_' + serviceName + '"' in cLst[i] and "create" not in cLst[i]:
            # print(i)
            p = i

            for j in range(p, 0, -1):
                if "entry " in cLst[j]:
                    # print(j)
                    for t in range(j, p + 2):
                        tmpLst.append(cLst[t])
                    break
            retLst.append(tmpLst)
            tmpLst = []

    return retLst


def getTheDeleteEntryId(tup, cfgLst):
    entryList = []
    expression_1 = ""
    serviceAddressStr = ""
    servicePortStr = ""
    if tup[7] != None:
        expression_str = tup[7].replace("http://", "")
        if ":" in expression_str and ":*" not in expression_str:
            expression_1 = expression_str.split(":")[0]
        else:
            if "/*" in expression_str:
                expression_str = expression_str.replace("/*", "")

            try:
                expression_1 = expression_str[0:expression_str.index("/")]
            except:
                expression_1 = expression_str

    if tup[4] != None:
        if "/" not in tup[4]:
            ipstr = tup[4] + "/32"
        else:
            ipstr = tup[4]
        serviceAddressStr = ipstr

    if tup[6] != None:
        servicePortStr = str(tup[6])

    entryList = getServiceEntryList(tup[3], cfgLst)
    delete_entry_id = -1

    for entry in entryList:
        # print(entry)
        bl = judgeTheDeleteEntry(expression_1, serviceAddressStr, servicePortStr, entry)
        if bl == True:
            delete_entry_id = int(entry[0].split("entry ")[1].split(" ")[0])
            return delete_entry_id

    return delete_entry_id


def judgeTheDeleteEntry(express, serviceAddress, servicePort, ety):
    if express == "":
        for line in ety:
            if "expression 1" in line:
                return False
    else:
        for line in ety:
            if "expression 1" in line and express not in line:
                return False

    if serviceAddress == "":
        for line in ety:
            if "server-address" in line:
                return False
    else:
        for line in ety:
            if "server-address" in line and serviceAddress not in line:
                return False

    if servicePort == "":
        for line in ety:
            if "server-port" in line:
                return False
    else:
        for line in ety:
            if "server-port" in line and servicePort not in line:
                return False

    return True


def LimitTheServiceSpeed(comLst, tup):
    pass


def DeleteTheEntry(comLst, tup, entry_id):
    # print(entry_id)
    global serviceEntryIdDict
    comLst.append("exit all\n")
    comLst.append("configure application-assurance group 1:1 policy\n")
    comLst.append("begin\n")
    comLst.append("no entry " + str(entry_id))
    comLst.append("\n")
    print("移除的ID", entry_id, "移除前：", serviceEntryIdDict[tup[3]])
    try:
        serviceEntryIdDict[tup[3]].remove(entry_id)
    except:
        pass
    print("移除后：", serviceEntryIdDict[tup[3]])


def addEntryIdtoserviceEntryIdDict(serviceName, cfglst):
    global serviceEntryIdDict
    entryIdList = []
    if serviceName in serviceEntryIdDict:
        return ""

    for i in range(0, len(cfglst)):
        if 'application "APP_' + serviceName + '"' in cfglst[i] and "create" not in cfglst[i]:
            k = i
            # print(serviceName,k)
            for j in range(k, 0, -1):
                if 'server-address eq ip-prefix-list "app_' + serviceName in cfglst[j]:
                    break
                if "entry" in cfglst[j]:
                    # print(cfglst[j])
                    entryIdList.append(int(cfglst[j].split("entry ")[1].split(" create")[0]))
                    break
    serviceEntryIdDict[serviceName] = entryIdList
    # print(serviceName,entryIdList)


def getAllEntryIdList(all_entry_list, cfglst):
    for i in range(0, len(cfglst)):
        if 'entry' in cfglst[i] and "create" in cfglst[i]:
            all_entry_list.append(int(cfglst[i].split("entry ")[1].split(" create")[0]))


def setPRUCRUtoServiceDict(tup, cfglst):
    global serviceDict
    serviceId = tup[2]
    serviceName = tup[3]
    if "PR_" + serviceName not in serviceDict:
        serviceDict["PR_" + serviceName] = False
    if "PRU_" + serviceName + "_" + tup[0] not in serviceDict:
        serviceDict["PRU_" + serviceName + "_" + tup[0]] = False
    if "CRU_" + serviceName not in serviceDict:
        serviceDict["CRU_" + serviceName] = False

    # 判断PR是否存在
    prStr = 'policy-rule "PR_' + serviceName + '"'
    if PR_str_isExist(prStr, cfglst) == True:
        serviceDict["PR_" + serviceName] = True

    # 判断PRU是否存在
    pruStr = 'policy-rule-unit "PRU_' + serviceName + '_' + tup[0] + '"'
    if PRU_str_isExist(pruStr, cfglst) == True:
        serviceDict["PRU_" + serviceName + "_" + tup[0]] = True
        # print(serviceName+"_PRU",serviceDict[serviceName+"_PRU"])

    # 判断CRU是否存在
    cruStr = 'charging-rule-unit "CRU_' + serviceName + '"'
    if CRU_str_isExist(cruStr, serviceId, cfglst) == True:
        serviceDict["CRU_" + serviceName] = True


def PR_str_isExist(pr_str, cfglst):
    i = 0
    for text in cfglst:
        if pr_str in text and "qci * arp * precedence" not in text:
            i += 1
    if i == 2:
        return True
    else:
        return False


def PRU_str_isExist(pru_str, cfglst):
    for text in cfglst:
        if pru_str in text and "qci * arp * precedence" not in text:
            return True
    return False


def CRU_str_isExist(cru_str, sid, cfglst):
    global commandList

    for i in range(0, len(cfglst)):
        if cru_str in cfglst[i] and "qci * arp * precedence" not in cfglst[i]:
            # print(i,cfglst[i+1])
            try:
                if cfglst[i + 1].split("rating-group ")[1].replace("\n", "") != str(sid):
                    # print(cfglst[i+1].split("rating-group ")[1].replace("\n",""))
                    print(cru_str + "该ID：" + str(sid) + "匹配不对")
                    commandList.append("注意\n")
                    commandList.append(cru_str + "该ID：" + str(sid) + "匹配不对\n")
            except:
                pass
            return True


def PRU_CRU_is_Associate(serviceName, prStr, pruStr, clst):
    # print(serviceName)
    # print(pruStr)
    prStr = 'policy-rule "' + prStr + '"'
    prustr = 'policy-rule-unit "' + pruStr + '"'
    cruStr = 'charging-rule-unit "CRU_' + serviceName + '"'
    # print("+++++",prStr,prustr,cruStr)
    for text in clst:
        if prStr in text and prustr in text and cruStr in text:
            # print(True)
            return True
    # print(False)
    return False


def PR_PRU_CRU_Process(lst, tup, cfglst):
    # print(tup)
    # 检测CRU是否存在，不存在则创建
    if serviceDict["CRU_" + tup[3]] == False:
        # 创建CRU
        lst.append('exit all' + "\n")
        lst.append("configure mobile-gateway profile policy-options " + "\n")
        lst.append('charging-rule-unit "CRU_' + tup[3] + '" ' + "\n")
        lst.append('rating-group ' + str(tup[2]) + "\n")
        lst.append('service-identifier ' + str(tup[2]) + "\n")
        lst.append('reporting-level service-id' + "\n")
        lst.append('exit' + "\n")
        lst.append("\n")
        serviceDict["CRU_" + tup[3]] = True

    '''
    #检测PR是否存在，不存在则创建
    if serviceDict["PR_" + tup[3]] == False:
        lst.append("该业务需要创建PR\n")
        lst.append('exit all' + "\n")
        lst.append("configure mobile-gateway profile policy-options " + "\n")
        lst.append("该创建PR的命令\n")
        lst.append("\n")
        serviceDict["PR_" + tup[3]] = True
    '''

    # 检测PRU是否存在
    if serviceDict["PRU_" + tup[3] + '_' + tup[0]] == False:
        lst.append('exit all' + "\n")
        lst.append("configure mobile-gateway profile policy-options " + "\n")
        lst.append("begin" + "\n")
        pruKey = "PRU_" + tup[3] + '_' + tup[0]
        # pruStr = 'policy-rule-unit "' + pruKey + '" create' + "\n"
        pruStr = 'policy-rule-unit "' + pruKey + "\n"
        lst.append(pruStr)
        lst.append('flow-description ' + str(1) + "\n")
        lst.append('match' + "\n")
        lst.append('aa-charging-group "CHG_' + tup[3] + '"' + "\n")
        lst.append('exit' + "\n")
        lst.append('exit' + "\n")
        lst.append("\n")
        serviceDict["PRU_" + tup[3] + '_' + tup[0]] = True

    # 检测PRU,CRU是否关联到PR
    if serviceDict["PR_" + tup[3]] == False:
        serviceDict["PR_" + tup[3]] = True
        # if PRU_CRU_is_Associate(tup,cfglst) == False:
        #print(tup[3], "需要关联PR,PRU,CRU")
        precedenceId = 10000
        pruKey = "PRU_" + tup[3] + '_' + tup[0]
        pruStr = 'policy-rule-unit "' + pruKey + '"'
        cmpstr = 'policy-rule "PR_' + tup[3] + '" ' + pruStr + ' charging-rule-unit "CRU_' + tup[
            3] + '" qci * arp * precedence ' + str(precedenceId)
        lst.append(cmpstr)
        lst.append('\n')


def PRU_CRU_is_Associate(tup, clst):
    prStr = 'policy-rule "PR_' + tup[3] + '"'
    prustr = 'policy-rule-unit "PRU_' + tup[3] + '_' + tup[0] + '"'
    cruStr = 'charging-rule-unit "CRU_' + tup[3] + '"'
    # print("+++++",prStr,prustr,cruStr)
    for text in clst:
        if prStr in text and prustr in text and cruStr in text:
            # print(True)
            return True
    # print(False)
    return False


def PR_PRU_CRU_Delete(lst, tup, cfglst):
    # 检测该字典中(serviceEntryIdDict)业务的entryId 列表是否为空，若空则表示没改业务了需要删PR,PRU,CRU等
    global serviceEntryIdDict
    if len(serviceEntryIdDict[tup[3]]) == 0:
        print(tup[3], "该业务已经删完了,需要删除相应的PRU,CRU,PR,以及删除关联")
        lst.append(tup[3] + "该业务已经删完了,需要删除相应的PRU,CRU,PR,以及删除关联\n")
